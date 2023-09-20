import geopandas as gpd
import pandas as pd
import ast
import warnings
import time
import numpy as np
from fiona.drvsupport import supported_drivers
from openpyxl import load_workbook

from Costeo import costeototal
import matplotlib.pyplot as plt
import math
import os
os.environ['USE_PYGEOS'] = '0'
supported_drivers['LIBKML'] = 'rw'
warnings.filterwarnings('ignore')



#Funciones
def ponerfecha(row):
    import datetime
    a = datetime.datetime(2023, 9, 10, row['HORA'], row['MINUTO'])
    return a


start = time.time()

# Parametros
intervalofrecuencia = 15 # minutos
capacidadcoche = 40  # personas promedio sobre el coche
velocidadcaminando = 5 # km/h
longitudturno = 1 # hora
tolerancia = 1 # minutos
zona = "VECINAL"
codigoeg = "SFE_GTRXV35_PE5000h8cortado_Base_Sindiez"
codigoef = "SFE_FA35_AC_BaseCR_"
codigos = "SFE_TFAV2_AC_BaseCR_"
sistemanuevo = False
directorio = "CarrilRapido/"
TipoDeTuning = "mincoches" # "cercania" o "minfreq" o "mincoches"
graficar = False


# Direccion entradas
direccionlineaNueva = "SHP Files/Nuevas Lineas/N.kml"  # modificado # agregué la línea superposición 16
direccionlineas = "SHP Files/Nuevas Lineas/Lineas Actuales.kml"  # direccion de las lineas
direccionrutas = "Output/Pruebas Nuevos Sistemas/" + directorio + codigoeg + "RutasCompletas.csv"  # direccion de las rutas
direccionparadas = "Output/Pruebas Nuevos Sistemas/" + directorio + codigoeg + "ParadasPromedio.csv"
direccionnodos = "Output/Pruebas Nuevos Sistemas/" + directorio + codigoeg + "nodos.xlsx"
direccionzonas = "SHP Files/Vecinales/Vecinales.shp"
direcciongeolocalizacion = "Preprocessdata/paradasVF.xlsx"  # direccion de las paradas
direccionfa = "Output/Pruebas Nuevos Sistemas/" + directorio + codigoef + "FrecuenciasDesagregadas.xlsx"
direcciontablatuning = "Output/Pruebas Nuevos Sistemas/" + directorio + codigoef + "TablaTuning.xlsx"
direccioncargaT = "Output/Pruebas Nuevos Sistemas/" + directorio + codigoef + "CargaT.csv"
direccionmetricas = "Output/Pruebas Nuevos Sistemas/" + directorio + codigoeg + "Metricas.xlsx"
direccionpmporhora = "Output/Pruebas Nuevos Sistemas/" + directorio + codigoeg + "CargaPMPorHora.csv"
direccionmatrizlegado = "Output/Pruebas Nuevos Sistemas/Legado/SFE_TFAV1_LE_MatrizZonaaZona.xlsx" # es fijo
direccionperfilvelocidades = "Output/Pruebas Nuevos Sistemas/" + directorio + codigoef + "PerfilVelocidades.xlsx"
direccioncostos = 'Preprocessdata/Costospreprocesados.xlsx'
direccionparametros = 'Preprocessdata/ParametrosCosteo.xlsx' # modificado csv a xlsx


# Direccion salidas
directoriosalida = "Output/Pruebas Nuevos Sistemas/" + directorio + codigos


# pre procesamiento
RutasCompletas = pd.read_csv(direccionrutas)
fa = pd.read_excel(direccionfa)
SHP = gpd.read_file(direccionlineas, driver="LIBKML")  # cargo el shp de las lineas
SHPNueva = gpd.read_file(direccionlineaNueva)  # cargo el shp de la linea nueva # modificado
SHP = pd.concat([SHP,SHPNueva],axis=0)  # concateno ambos gdfs
SHP = SHP.rename(columns={'Name': 'Lineas'})
SHP = SHP.to_crs(epsg=22185)

Lineas = SHP["Lineas"].unique()
SHP = SHP[['Lineas', 'geometry']]
CargaT = pd.read_csv(direccioncargaT)

for i, row in SHP.iterrows():
    length = row['geometry'].length
    SHP.at[i, 'Longitud'] = length
nodos = pd.read_excel(direccionnodos)
perfilvelocidades = pd.read_excel(direccionperfilvelocidades)

TablaTuning = pd.read_excel(direcciontablatuning)
TablaTuning["FrecuenciaPropuesta"] = intervalofrecuencia #modificado agregado por ale porque no estaba como input
TablaTuning["Factor"] = TablaTuning["Frecuencia"] / TablaTuning["FrecuenciaPropuesta"]
TablaTuning["Capacidad"] = capacidadcoche
TablaTuning["CapacidadCorregida"] = TablaTuning["Capacidad"] / TablaTuning["Factor"]
TablaTuning = TablaTuning[["Linea", "CapacidadCorregida", "InicioTurno", "FrecuenciaPropuesta", "Factor"]]



# procesamiento
grupoporhora = fa.groupby([pd.Grouper(key='inicio', freq=str(longitudturno) + "H", offset="5h"), "LineaUtilizada"])[
    "0"].max().reset_index()
grupoporhora = pd.merge(grupoporhora, SHP, left_on="LineaUtilizada", right_on="Lineas", how="left")
grupoporhora["HORA"] = grupoporhora["inicio"].dt.hour
grupoporhora = grupoporhora[['inicio', 'Lineas', '0', 'Longitud', "HORA"]]
grupoporhora = pd.merge(grupoporhora, perfilvelocidades, left_on=["Lineas", "HORA"], right_on=["Lineas", "HORA"], how="left")
grupoporhora = pd.merge(grupoporhora, TablaTuning, left_on=["Lineas", "inicio"], right_on=["Linea", "InicioTurno"], how="left")
grupoporhora = grupoporhora[["inicio", "Lineas", "0", "Longitud", "Velocidad", "CapacidadCorregida", "FrecuenciaPropuesta", "Factor"]]
grupoporhora["Tiempoporvuelta"] = (grupoporhora["Longitud"] / (grupoporhora["Velocidad"] * 1000))

if TipoDeTuning == "minfreq":
    grupoporhora["Coches"] = np.ceil((grupoporhora["Tiempoporvuelta"] / grupoporhora["FrecuenciaPropuesta"]) * 60)
    grupoporhora["Coches"] = grupoporhora["Coches"].replace([np.inf, -np.inf], 0)
    grupoporhora["FrecuenciaImplicita"] = (grupoporhora["Tiempoporvuelta"] / grupoporhora["Coches"]) * 60
    grupoporhora["FrecuenciaImplicita"] = grupoporhora["FrecuenciaImplicita"].replace([np.inf, -np.inf], 0)
elif TipoDeTuning == "cercania":
    Posibilidades = []
    lineastv = grupoporhora[["Lineas", "Tiempoporvuelta", "inicio"]].drop_duplicates()
    for i, tv in enumerate(lineastv["Tiempoporvuelta"]):
        for c in range(1, 25):
            frecuencia = (tv / c) * 60
            dictinterno = {"Linea": lineastv.at[i, "Lineas"],
                           "Coches": c, "inicio" : lineastv.at[i, "inicio"],
                           "FrecuenciaImplicita": frecuencia}
            Posibilidades += [dictinterno]
    Posibilidades = pd.DataFrame(Posibilidades)
    Posibilidades = pd.merge(Posibilidades, TablaTuning, left_on=["Linea", "inicio"], right_on=["Linea", "InicioTurno"], how="left")
    Posibilidades["Diferencia"] = abs(Posibilidades["FrecuenciaImplicita"] - Posibilidades["FrecuenciaPropuesta"])
    Pseleccion = Posibilidades.groupby(["Linea", "inicio"])["Diferencia"].min().reset_index()
    Pseleccion = pd.merge(Pseleccion, Posibilidades, left_on=["Linea", "inicio", "Diferencia"], right_on=["Linea", "inicio", "Diferencia"], how="inner")
    Pseleccion = Pseleccion[["Linea", "inicio", "Coches", "FrecuenciaImplicita"]]
    grupoporhora = pd.merge(grupoporhora, Pseleccion, left_on=["Lineas", "inicio"], right_on=["Linea", "inicio"], how="inner").dropna()
else:
    grupoporhora["Coches"] = np.floor((grupoporhora["Tiempoporvuelta"] / grupoporhora["FrecuenciaPropuesta"]) * 60)
    grupoporhora["Coches"] = grupoporhora["Coches"].replace([np.inf, -np.inf], 0)
    grupoporhora["FrecuenciaImplicita"] = (grupoporhora["Tiempoporvuelta"] / grupoporhora["Coches"]) * 60
    grupoporhora["FrecuenciaImplicita"] = grupoporhora["FrecuenciaImplicita"].replace([np.inf, -np.inf], 0)

grupoporhora["kmrecorridos"] = (longitudturno / grupoporhora["Tiempoporvuelta"]) * grupoporhora["Coches"] * grupoporhora[
    "Longitud"] / 1000
cochesporturno = grupoporhora.groupby(["inicio"])["Coches"].sum().reset_index()
cochesporturno.rename(columns={"Coches": "TotalCochesTurno"}, inplace=True)
distanciarecorridoporlinea = grupoporhora.groupby(["Lineas"])["kmrecorridos"].sum().reset_index()
grupoporhora["Utilización"] = grupoporhora["0"] * 100 / (capacidadcoche * grupoporhora["Coches"] / grupoporhora["Tiempoporvuelta"])


Ocio = fa.groupby(["LineaUtilizada", pd.Grouper(key='inicio', freq=str(longitudturno) + "H")])["0"].max().reset_index()
Ocio["HORA"] = Ocio["inicio"].dt.hour
Ocio.rename(columns={"0": "CargaMaxima"}, inplace=True)
fa["HORA"] = fa["inicio"].dt.hour
Ocio = pd.merge(fa, Ocio, on=["LineaUtilizada", "HORA"], how="left")
Ocio["Ocupacion"] = Ocio["0"] / Ocio["CargaMaxima"]
OcioPromedio = Ocio.groupby(["LineaUtilizada", pd.Grouper(key='inicio_x', freq=str(longitudturno) + "H", offset="5h")])[
    "Ocupacion"].mean().reset_index()
OcioPromedio["Ocupacion"] = OcioPromedio["Ocupacion"] * 100


pmporhora = pd.read_csv(direccionpmporhora)
pmporhora["MINUTO"] = 0
pmporhora['FECHATRX'] = pmporhora.apply(lambda row: ponerfecha(row), axis=1)
pmporhora[['LineaUtilizada', 'Orden']] = pmporhora["Nodo"].str.split("-", n=1, expand=True)
pmax = pmporhora.groupby(["LineaUtilizada", pd.Grouper(key='FECHATRX', freq=str(longitudturno) + "H", offset = "5H")])["Personas"].max().reset_index()
pmax = pmax.rename(columns={"Personas": "MaximoPersonas"})
Homogeneidad = pd.merge(pmporhora, pmax, on=["LineaUtilizada", "FECHATRX"], how="left")
Homogeneidad["Proporcion"] = Homogeneidad["Personas"] / Homogeneidad["MaximoPersonas"]
HomogeneidadTurno = Homogeneidad.groupby(["LineaUtilizada", pd.Grouper(key='FECHATRX', freq=str(longitudturno) + "H", offset = "5H")])["Proporcion"].mean().reset_index()
HomogeneidadTurno.rename(columns={"Proporcion": "Homogeneidad"}, inplace=True)
HomogeneidadPromedio = HomogeneidadTurno.Homogeneidad.mean()*100

CargaT["FECHATRX"] = pd.to_datetime(CargaT["FECHATRX"])
Cargaporlinea = CargaT.groupby(["LineaUtilizada", pd.Grouper(key='FECHATRX', freq=str(longitudturno) + "H", offset="5h")])[
    "CargaProporcional"].sum().reset_index()
Cargaporlinea = Cargaporlinea[Cargaporlinea["LineaUtilizada"].isin(Lineas)]


# Medias

MediaPM = fa.groupby(["index"])["0"].mean().reset_index()
MediaPM["Separar"] = MediaPM["index"]
MediaPM[["LineaUtilizada", "Orden"]] = MediaPM["Separar"].str.split("-", n=1, expand=True)
MediaPM["Orden"] = MediaPM["Orden"].astype(float)
MediaPM.sort_values(by=["LineaUtilizada", "Orden"], inplace=True)
MediaPM.rename(columns={"0": "Media"}, inplace=True)

DesvioPM = fa.groupby(["index"])["0"].std().reset_index()
DesvioPM["Separar"] = DesvioPM["index"]
DesvioPM[["LineaUtilizada", "Orden"]] = DesvioPM["Separar"].str.split("-", n=1, expand=True)
DesvioPM["Orden"] = DesvioPM["Orden"].astype(float)
DesvioPM.rename(columns={"0": "Desvio"}, inplace=True)
DesvioPM.sort_values(by=["LineaUtilizada", "Orden"], inplace=True)

MediaPM = pd.merge(MediaPM, DesvioPM, on=["index", "LineaUtilizada", "Orden"], how="inner")
MediaPM.to_excel(directoriosalida + "MediaPM.xlsx", index=False)


# Tiempos de viaje

RutasSimples = RutasCompletas[(RutasCompletas["NroLineas"] == 1) & (RutasCompletas["LineaUtilizada"] != "Ninguna")][["Nodo Origen", "Nodo Destino", "DistanciaCaminando", "LineaUtilizada"]]
RutasCombinadas = RutasCompletas[RutasCompletas["NroLineas"] > 1][["Nodo Origen", "Nodo Destino", "DistanciaCaminando", "LineaUtilizada"]]

RutasSimples["TiempoCaminando"] = (RutasSimples["DistanciaCaminando"] / velocidadcaminando) * (60/1000)
TiempoDeViaje = CargaT[['Nodo Origen', 'Nodo Destino', 'LineaUtilizada', "RutaT","HORA"]] # modificado agregado "inicio"
TiempoDeViaje.drop_duplicates(subset=["Nodo Origen", "Nodo Destino", "LineaUtilizada"], inplace=True)
TiempoDeViaje["RutaT"] = TiempoDeViaje["RutaT"].astype(str)
TiempoDeViaje['RutaT'] = TiempoDeViaje['RutaT'].apply(ast.literal_eval)
TiempoDeViaje["TiempoDeViaje"] = TiempoDeViaje["RutaT"].apply(lambda x: x[-1])
TiempoDeViaje["TiempoDeViaje"] = TiempoDeViaje["TiempoDeViaje"].astype(float)
TiempoDeViajePromxLinea = TiempoDeViaje.groupby(["LineaUtilizada","HORA"])["TiempoDeViaje"].mean().reset_index() # modificado para categorizar
TiempoDeViajePromxLinea["MINUTO"] = 0 # modificado # agregado de columna minutos 0
TiempoDeViajePromxLinea["InicioTurno"] = TiempoDeViajePromxLinea.apply(lambda x: ponerfecha(x),axis=1) # modificado para categorizar
TiempoDeViajePromxLinea = TiempoDeViajePromxLinea[["InicioTurno","LineaUtilizada","TiempoDeViaje"]] # modificado para categorizar
TiempoDeViaje = TiempoDeViaje[["Nodo Origen", "Nodo Destino", "LineaUtilizada", "TiempoDeViaje"]]

TiempoDeEspera = grupoporhora[["inicio", "Lineas", "FrecuenciaImplicita"]]
TiempoDeEspera["TiempoDeEspera"] = TiempoDeEspera["FrecuenciaImplicita"]/2

tiemposdecorte = TiempoDeEspera["inicio"].unique()
TiempoDeEspera["inicio"] = TiempoDeEspera["inicio"].astype('datetime64[ns]')
fechaextra = tiemposdecorte[-1] + pd.to_timedelta(longitudturno+0.5, unit='h')
fechaextra = np.array(fechaextra)
tiemposdecorte = np.append(tiemposdecorte, fechaextra).astype('datetime64[ns]')
intervalos = pd.IntervalIndex.from_breaks(breaks=tiemposdecorte, closed='right')
CargaT['intervalo'] = pd.cut(CargaT['FECHATRX'], bins=intervalos, labels=range(len(intervalos)))
CargaT['inicio'] = CargaT["intervalo"].apply(lambda x: x.left)
CargaTparatiempos = CargaT[["Nodo Origen", "Nodo Destino", "LineaUtilizada", "inicio", "CargaProporcional"]]
CargaTparatiempos["inicio"] = CargaTparatiempos["inicio"].astype('datetime64[ns]')
CargaTparatiempos = pd.merge(CargaTparatiempos, TiempoDeEspera, left_on=["inicio", "LineaUtilizada"], right_on=["inicio", "Lineas"], how="left")
CargaTparatiempos = pd.merge(CargaTparatiempos, TiempoDeViaje, left_on=["Nodo Origen", "Nodo Destino", "LineaUtilizada"], right_on=["Nodo Origen", "Nodo Destino", "LineaUtilizada"], how="left")
CargaTparatiempos = pd.merge(CargaTparatiempos, RutasSimples, left_on=["Nodo Origen", "Nodo Destino", "LineaUtilizada"], right_on=["Nodo Origen", "Nodo Destino", "LineaUtilizada"], how="left")
CargaTparatiempos = CargaTparatiempos[['Nodo Origen', 'Nodo Destino', 'LineaUtilizada', 'inicio', 'TiempoDeEspera', 'TiempoDeViaje', 'TiempoCaminando', 'CargaProporcional']]
CargaTparatiempos["TiempoTotal"] = CargaTparatiempos["TiempoDeEspera"] + CargaTparatiempos["TiempoDeViaje"] + CargaTparatiempos["TiempoCaminando"]
CargaSumada = CargaTparatiempos["CargaProporcional"].sum()
CargaTparatiempos["TiempoPonderado"] = CargaTparatiempos["TiempoTotal"] * CargaTparatiempos["CargaProporcional"] / CargaSumada
CargaTparatiempos["TiempoPonderadoC"] = CargaTparatiempos["TiempoCaminando"] * CargaTparatiempos["CargaProporcional"] / CargaSumada
CaminataPromedio = CargaTparatiempos.TiempoPonderadoC.sum()
TiempoPromedio = CargaTparatiempos.TiempoPonderado.sum()
CargaTparatiempos.to_csv(directoriosalida + "TiemposDesagregados.csv", index=False)

# IPK # modificado corrido todo este bloque más abajo
IPK = pd.merge(Cargaporlinea, grupoporhora, left_on=["LineaUtilizada", "FECHATRX"], right_on=["Lineas", "inicio"],
               how="inner")
IPK["IPK"] = IPK["CargaProporcional"] / IPK["kmrecorridos"]
IPK = IPK[["LineaUtilizada", "inicio", "CargaProporcional", "kmrecorridos", "Coches", "FrecuenciaImplicita", "IPK", "Utilización"]]
IPK = pd.merge(IPK, cochesporturno, left_on="inicio", right_on="inicio", how="left")
IPK.replace([np.inf, -np.inf], np.nan, inplace=True)
IPK = IPK.dropna()
IPK.drop(columns=["TotalCochesTurno"], inplace=True)
IPK = pd.merge(IPK, OcioPromedio, left_on=["LineaUtilizada", "inicio"], right_on=["LineaUtilizada", "inicio_x"],
               how="left")

IPK.rename(columns={"inicio": "InicioTurno", "LineaUtilizada": "Linea", "CargaProporcional": "TRXTurno",
                    "kmrecorridos": "DistanciaRecorrida", "Coches": "NroVehiculos",
                    "FrecuenciaImplicita": "Frecuencia"}, inplace=True)
IPK = pd.merge(IPK, SHP, left_on="Linea", right_on="Lineas", how="left")
IPK = pd.merge(IPK, HomogeneidadTurno, left_on=["Linea", "InicioTurno"], right_on=["LineaUtilizada", "FECHATRX"], how="left")
IPK = IPK[["Linea", "InicioTurno", "TRXTurno", "NroVehiculos", "DistanciaRecorrida", "Longitud", "Frecuencia", "Ocupacion", "Utilización", "Homogeneidad", "IPK"]]
IPK = pd.merge(IPK,TiempoDeViajePromxLinea,left_on=["Linea","InicioTurno"],right_on=["LineaUtilizada","InicioTurno"],how='left').drop(["LineaUtilizada"],axis=1) # modificado agregado TPO DE VIAJE
Costos = costeototal(direccioncostos, direccionparametros, IPK, directoriosalida+"Costos.xlsx", "dia", 120)[1]
IPK = pd.merge(IPK, Costos, left_on=["Linea"], right_on=["Línea"], how="left")

IPKpromedio = IPK.IPK.mean()
NroCochesMinimos = IPK.groupby(["InicioTurno"])["NroVehiculos"].sum().mean()
FrecuenciaPromedio = IPK.groupby(["InicioTurno"])["Frecuencia"].mean().mean()
DistanciaDiaria = IPK["DistanciaRecorrida"].sum()
OcupacionPromedio = IPK["Ocupacion"].mean()
UtilizacionPromedio = IPK["Utilización"].mean()
IPK.to_excel(directoriosalida + "ReporteCompletoIPK.xlsx", index=False)

CostoTotal = IPK.groupby(["Linea"])["Costo"].mean().sum() #modificado # es mean porque el costo diario esta puesto en cada turno

#Resumen zona a zona
shpzona = gpd.read_file(direccionzonas).set_crs(epsg=4326)
paradas = pd.read_excel(direcciongeolocalizacion)
paradas["IDcompleto"] = paradas["ID"].astype(str) + "-" + paradas["Lineas"].astype(str)
puntosparadas = gpd.GeoDataFrame(paradas, geometry=gpd.points_from_xy(paradas.Longitud, paradas.Latitud), crs=4326)
puntosparadas = puntosparadas[["IDcompleto", "geometry"]]
puntosparadas = gpd.sjoin_nearest(puntosparadas, shpzona, how="left")
puntosparadas = puntosparadas[["IDcompleto", "geometry", zona]].drop_duplicates(subset=["IDcompleto"])
CargaTparatiempos = pd.merge(CargaTparatiempos, puntosparadas, left_on=["Nodo Origen"], right_on=["IDcompleto"], how="left")
CargaTparatiempos = pd.merge(CargaTparatiempos, puntosparadas, left_on=["Nodo Destino"], right_on=["IDcompleto"], how="left")
CargaTparatiempos["inicio"] = pd.to_datetime(CargaTparatiempos["inicio"])
CargaTparatiempos["HORA"] = CargaTparatiempos["inicio"].dt.hour
MatrizZonaaZonaTT = CargaTparatiempos.groupby([zona + "_x", zona + "_y", "HORA"])[["TiempoTotal"]].mean().reset_index()
CargaParaMatrizZonaaZona = CargaTparatiempos.groupby([zona + "_x", zona + "_y", "HORA"])[["CargaProporcional"]].sum().reset_index()
MatrizZonaaZonaTT = pd.merge(MatrizZonaaZonaTT, CargaParaMatrizZonaaZona, left_on=[zona + "_x", zona + "_y", "HORA"], right_on=[zona + "_x", zona + "_y", "HORA"], how="inner")
MatrizZonaaZonaTC = CargaTparatiempos.groupby([zona + "_x", zona + "_y", "HORA"])[["TiempoCaminando"]].mean().reset_index()
MatrizZonaaZonaTC = pd.merge(MatrizZonaaZonaTC, CargaParaMatrizZonaaZona, left_on=[zona + "_x", zona + "_y", "HORA"], right_on=[zona + "_x", zona + "_y", "HORA"], how="inner")

MatrizZonaaZonaTT.to_excel(directoriosalida + "MatrizZonaaZona.xlsx", index=False, sheet_name="TiempoTotal")
book = load_workbook(directoriosalida + "MatrizZonaaZona.xlsx")
writer = pd.ExcelWriter(directoriosalida + "MatrizZonaaZona.xlsx", engine = 'openpyxl')
writer.book = book

MatrizZonaaZonaTC.to_excel(writer, index=False, sheet_name="TiempoCaminando")
writer.close()



if sistemanuevo:
    matrizlegadott = pd.read_excel(direccionmatrizlegado, sheet_name="TiempoTotal")
    matrizlegadotc = pd.read_excel(direccionmatrizlegado, sheet_name="TiempoCaminando")
    matrizdiferenciatt = pd.merge(MatrizZonaaZonaTT, matrizlegadott, left_on=[zona + "_x", zona + "_y", "HORA"], right_on=[zona + "_x", zona + "_y", "HORA"], how="inner", suffixes=("_Nuevo", "_Legado"))
    matrizdiferenciatt["Diferencia"] = matrizdiferenciatt["TiempoTotal_Nuevo"] - matrizdiferenciatt["TiempoTotal_Legado"]
    matrizdiferenciatt["DiferenciaProporcional"] = matrizdiferenciatt["Diferencia"] * matrizdiferenciatt["CargaProporcional_Legado"]
    matrizdiferenciatc = pd.merge(MatrizZonaaZonaTC, matrizlegadotc, left_on=[zona + "_x", zona + "_y", "HORA"], right_on=[zona + "_x", zona + "_y", "HORA"], how="inner", suffixes=("_Nuevo", "_Legado"))
    matrizdiferenciatc["Diferencia"] = matrizdiferenciatc["TiempoCaminando_Nuevo"] - matrizdiferenciatc["TiempoCaminando_Legado"]
    matrizdiferenciatc["DiferenciaProporcional"] = matrizdiferenciatc["Diferencia"] * matrizdiferenciatc["CargaProporcional_Legado"]
    matrizdiferenciatt.to_excel(directoriosalida + "MatrizZonaaZonaDiferencia.xlsx", index=False, sheet_name="TiempoTotal")
    book = load_workbook(directoriosalida + "MatrizZonaaZonaDiferencia.xlsx")
    writer = pd.ExcelWriter(directoriosalida + "MatrizZonaaZonaDiferencia.xlsx", engine = 'openpyxl')
    writer.book = book
    matrizdiferenciatc.to_excel(writer, index=False, sheet_name="TiempoCaminando")
    writer.close()
    DiferenciaTT = matrizdiferenciatt["DiferenciaProporcional"].sum() / matrizdiferenciatt["CargaProporcional_Legado"].sum()
    DiferenciaTC = matrizdiferenciatc["DiferenciaProporcional"].sum() / matrizdiferenciatc["CargaProporcional_Legado"].sum()
else:
    DiferenciaTT = 0
    DiferenciaTC = 0




end = time.time()
tiempoejecucion = (end - start) / 60
print("Tiempo de ejecución: " + str(tiempoejecucion) + " minutos")

book = load_workbook(direccionmetricas)
writer = pd.ExcelWriter(direccionmetricas, engine = 'openpyxl')
writer.book = book

metricas = pd.DataFrame({
    'IPKp': [IPKpromedio],
    'NroCochesMinimo': [NroCochesMinimos], "FrecuenciaPromedio" : [FrecuenciaPromedio],
    'OcupacionPromedio%': [OcupacionPromedio],'UtilizacionPromedio%': [UtilizacionPromedio], "Homogeneidad%": [HomogeneidadPromedio], "KMTotales" : [DistanciaDiaria], "TiempoPromedio" : [TiempoPromedio], "TiempoCaminataPromedio" : [CaminataPromedio], "DiferenciaTiempoTotal" : [DiferenciaTT],  "DiferenciaTiempoCaminando" : [DiferenciaTC], "TiempoEjecucion": [tiempoejecucion], "TipoDeTuning" : [TipoDeTuning],
    'CostoTotal (ARS)' : [CostoTotal]
}).T

metricas.to_excel(writer, sheet_name = 'MetricasTuning', index=True, engine = 'openpyxl')
writer.close()




# graficos
if graficar:
    for l in grupoporhora.Lineas.unique():
        #Coches por hora
        datosg = grupoporhora[grupoporhora.Lineas == l].sort_values(by=["inicio"], ascending=True)
        datosg["HORA"] = datosg["inicio"].dt.hour
        fig, ax = plt.subplots(figsize=(16, 9))
        plt.plot(datosg['HORA'], datosg['Coches'], marker='o', linestyle='--')
        plt.xlabel('HORA')
        plt.ylabel('Coches')
        minimo = datosg['Coches'].min()
        maximo = datosg['Coches'].max()
        plt.axhline(y=maximo, color='r', linestyle='-')
        plt.axhline(y=minimo, color='r', linestyle='-')
        ax.grid( color='green',
                linestyle='-.', linewidth=0.5,
                alpha=0.2)
        ax.set_title('Variación en la necesidad de moviles para la linea ' + l,
                     loc='left', fontname="Gotham Rounded", fontsize=20, pad=10, color="dimgrey")
        plt.suptitle(
            "Coches por hora",
            fontsize=35,
            fontweight="bold", fontname="Gotham Rounded",
            x=0.124,
            y=0.975,
            ha="left")
        for tick in ax.get_xticklabels():
            tick.set_fontname("Lato")
            tick.set_fontsize("13")
        for tick in ax.get_yticklabels():
            tick.set_fontname("Lato")
            tick.set_fontsize("13")
        plt.annotate('Fuente: Algoritmo TFAV', (0, 0), (0, -30), fontsize=9,
                     xycoords='axes fraction', textcoords='offset points', va='top')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        plt.savefig(directoriosalida + "Cochesporhora" + l + ".png", dpi=400, bbox_inches='tight')
        plt.close()

        #Utilización por hora

        fig, ax = plt.subplots(figsize=(16, 9))
        plt.bar(datosg['HORA'], datosg['Utilización'], color='#12bc8e', alpha=0.5)
        plt.xlabel('HORA')
        plt.ylabel('Utilización')
        minimo = datosg['Utilización'].min()
        maximo = datosg['Utilización'].max()
        plt.axhline(y=maximo, color='b', linestyle='-')
        plt.axhline(y=minimo, color='b', linestyle='-')
        plt.axhline(y=capacidadcoche, color='r', linestyle='-')
        plt.text(22, 41, 'Capacidad', fontsize=30, va='center', ha='left', backgroundcolor='w')

        ax.grid( color='green',
                linestyle='-.', linewidth=0.5,
                alpha=0.2)
        ax.set_title('Variación en la utilización de moviles para la linea ' + l,
                     loc='left', fontname="Gotham Rounded", fontsize=20, pad=10, color="dimgrey")
        plt.suptitle(
            "Utilización por hora",
            fontsize=35,
            fontweight="bold", fontname="Gotham Rounded",
            x=0.124,
            y=0.975,
            ha="left")
        for tick in ax.get_xticklabels():
            tick.set_fontname("Lato")
            tick.set_fontsize("13")
        for tick in ax.get_yticklabels():
            tick.set_fontname("Lato")
            tick.set_fontsize("13")
        plt.annotate('Fuente: Algoritmo TFAV', (0, 0), (0, -30), fontsize=9,
                     xycoords='axes fraction', textcoords='offset points', va='top')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        plt.savefig(directoriosalida + "Utilización" + l + ".png", dpi=400, bbox_inches='tight')
        plt.close()
else:
    print("No se graficaron las métricas")
